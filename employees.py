#importing openpyxl library to work with Excel files (.xlsx)
import openpyxl
# import datetime lib to work with dates.
from datetime import datetime

def read_excel_file(file_path):
    try:
        # Loading the workbook
        workbook = openpyxl.load_workbook(file_path)

        # gets the active sheet from the loaded workbook.(by default the active sheet in first page)
        sheet = workbook.active

        # set of names to store names of employees who has worked for 7 consecutive days
        names1 = set()

        # set of names of people who have less than 10 hours of time between shifts but greater than 1 hour
        names2 = set()

        # set of names of people who has worked for more than 14 hours in a single shift
        names3 = set()

        # Extract header row values (first row in the sheet)
        all_rows = list(sheet.iter_rows(values_only=True))
        header_row = all_rows[0]

        # defining column names as per the given spread sheet to get the index values
        name_col = "Employee Name"
        posid_col = "Position ID"
        time_col = "Time"
        timeout_col = "Time Out"
        shifthours_col = "Timecard Hours (as Time)"

        # getting index values of columns
        name_index = header_row.index(name_col) if name_col in header_row else None
        posid_index = header_row.index(posid_col) if posid_col in header_row else None
        time_index = header_row.index(time_col) if time_col in header_row else None
        timeout_index = header_row.index(timeout_col) if timeout_col in header_row else None
        shifthours_index = header_row.index(shifthours_col) if shifthours_col in header_row else None

        # defining variables
        prev_posid = ""
        prev_day = ""
        prev_time = ""
        day_count = 0

        # looping through each row
        for row in sheet.iter_rows(min_row=2):

            # extracting the values of required columns using index values.
            name = row[name_index].value
            posid = row[posid_index].value
            time = row[time_index].value
            timeout = row[timeout_index].value
            shifthours = row[shifthours_index].value
            total_hours = 0

            # skipping to next row if time-in or time-out are missing
            if time == "" or timeout == "":
                continue
            
            # extracting shift hours to check for any shifthours greater than 14
            if shifthours != "":
                striped_time = datetime.strptime(shifthours, '%H:%M')

                # Extract the hour and minute components
                shift_hours = striped_time.hour
                shift_minutes = striped_time.minute

                # Convert the time to hours (total hours, including the minute part)
                total_hours = shift_hours + shift_minutes / 60

            # checking for shifthours if greater than 14 hours
            if float(total_hours) > 14.00:
                names3.add(name)

            # extracting date of time-in
            if isinstance(time, datetime):
                time_date = time.date()
            
            #logic to find if any employer worked for 7 consecutive days
            if(prev_posid != posid):
                day_count = 1
            else:
                time_diff = (time_date - prev_day).days
                if(time_diff > 1):
                    day_count = 1
                elif(time_diff == 1):
                    day_count += 1
                else:
                    pass
            
            if(day_count == 7):
                names1.add(name)
            
            #logic to find if any employee have less than 10 hours of time between shifts but greater than 1 hour
            hours_diff = 0
            if(prev_posid == posid):
                time_diff = prev_time - time
                hours_diff = time_diff.seconds/3600

            if(hours_diff > 1 and hours_diff < 10):
                names2.add(name)

            # updating the variable
            prev_time = timeout
            prev_day = time_date
            prev_posid = posid

            

            
        # printing names of employees who has worked for 7 consecutive days
        print("Names of employees who has worked for 7 consecutive days:")
        for name in names1:
            print(name)
        print("")
        
        # printing names of employees who have less than 10 hours of time between shifts but greater than 1 hour
        print("Names of employees who have less than 10 hours of time between shifts but greater than 1 hour:")
        for name in names2:
            print(name)
        print("")

        # printing names of employees who has worked for more than 14 hours in a single shift
        print("Names of employees who has worked for more than 14 hours in a single shift:")
        for name in names3:
            print(name)

            

            

    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    # Get the file path from the user
    file_path = input("Enter the path to the .xlsx file: ")
    print("")
    # Call the function to read and print the contents
    read_excel_file(file_path)